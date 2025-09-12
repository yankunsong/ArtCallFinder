import os
import json
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

def write_to_excel(processed_data_dir='processed_data', output_file='art_calls.xlsx'):
    """
    Reads JSON files from a directory, processes the events, and writes them to an Excel file.

    Args:
        processed_data_dir (str): The directory containing the processed JSON files.
        output_file (str): The name of the output Excel file.
    """
    columns = ['reviewed', 'url', 'deadline', 'topics', 'fees', 'requirement', 'title', 'location', 'organization', 'source_file', 'added_on']
    
    existing_urls = set()
    df_existing = pd.DataFrame(columns=columns)
    
    # Check if the output file exists and read existing URLs
    if os.path.exists(output_file):
        try:
            df_existing = pd.read_excel(output_file, sheet_name='Sheet1')
            if 'url' in df_existing.columns:
                existing_urls = set(df_existing['url'])
        except Exception as e:
            print(f"Error reading existing Excel file: {e}")
    
    new_rows = []
    
    # Iterate over each file in the processed_data directory
    for filename in os.listdir(processed_data_dir):
        if filename.endswith('.json'):
            file_path = os.path.join(processed_data_dir, filename)
            
            with open(file_path, 'r') as f:
                try:
                    events = json.load(f)
                except json.JSONDecodeError:
                    print(f"Could not decode JSON from {filename}")
                    continue

            # Ensure events is a list
            if not isinstance(events, list):
                print(f"JSON file {filename} does not contain a list of events. Skipping.")
                continue

            for event in events:
                url = event.get('url')
                if url and url not in existing_urls:
                    new_row = {
                        'reviewed': "N",
                        'title': event.get('title'),
                        'deadline': event.get('deadline'),
                        'topics': ', '.join(event.get('topics_EN', [])),
                        'fees': event.get('fees'),
                        'requirement': event.get('requirement'),
                        'url': url,
                        'location': event.get('location'),
                        'organization': event.get('organization'),
                        'source_file': filename,
                        'added_on': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    new_rows.append(new_row)
                    existing_urls.add(url)

    if new_rows:
        df_new = pd.DataFrame(new_rows)
        
        # Ensure the columns are in the desired order
        df_new = df_new[columns]
        
        # Convert deadline to date format
        df_new['deadline'] = pd.to_datetime(df_new['deadline'], errors='coerce').dt.date

        if not os.path.exists(output_file):
            # File does not exist, create it
            with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
                df_new.to_excel(writer, index=False, sheet_name='Sheet1')
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']
                
                url_col_idx = df_new.columns.get_loc('url')
                url_format = workbook.add_format({'color': 'blue', 'underline': 1})
                
                for row_num, url in enumerate(df_new['url']):
                    if pd.notna(url):
                        worksheet.write_url(row_num + 1, url_col_idx, url, url_format, string=url)
            print(f"Created {output_file} and added {len(new_rows)} new events.")
        else:
            # File exists, append new data
            workbook = load_workbook(output_file)
            worksheet = workbook['Sheet1']
            
            # Get header and URL column index
            header = [cell.value for cell in worksheet[1]]
            try:
                url_col_idx = header.index('url')
            except ValueError:
                print("Error: 'url' column not found in Excel file.")
                return

            for _, row in df_new.iterrows():
                worksheet.append(row.tolist())
                
                # Apply hyperlink to URL
                new_row_num = worksheet.max_row
                url_cell = worksheet.cell(row=new_row_num, column=url_col_idx + 1)
                if url_cell.value:
                    url_cell.hyperlink = url_cell.value
                    url_cell.font = Font(color="0000FF", underline='single')

            workbook.save(output_file)
            print(f"Added {len(new_rows)} new events to {output_file}")
    else:
        print("No new events to add.")

if __name__ == '__main__':
    write_to_excel()
